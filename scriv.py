import enum
from functools import update_wrapper
from http.client import NETWORK_AUTHENTICATION_REQUIRED
from timeit import repeat
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill , numbers, Font, Border, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border, Side
import logging

class Scriv():
    def saveFile(self, fileName):
        logging.info("Saving file {}".format(fileName))

        print("Saving file...")
        for i in range(1, self.war.max_column + 1):
            self.war.column_dimensions[get_column_letter(i)].auto_size = True

        for i in range(1, self.capital.max_column + 1):
            self.capital.column_dimensions[get_column_letter(i)].auto_size = True

        for i in range(1, self.league.max_column + 1):
            self.league.column_dimensions[get_column_letter(i)].auto_size = True

        self.wb.save(fileName)

    def __init__(self, fileName, tags):
        self.raidDates = []
        self.raidTime = ""
        self.warDates = []
        self.warTime = ""
        self.seasons = []
        self.season = ""

        self.tags = tags

        self.clanData = {}
        self.clanMembers = {}

        self.outFlag = 0
        self.notAttackedFlag = 0

        self.numFormat = u'#,###;'

        self.thin_border = Border(top=Side(style='thin', color='454545'))
        self.topBorder = Border(top=Side(style='thick'))
        self.topNSideBorder = Border(top=Side(style='thick'), right=Side(style='medium'))
        self.bottomNSideBorder = Border(bottom=Side(style='thick'), right=Side(style='medium'))
        self.sideBorder = Border(right=Side(style='medium'))


        self.no_fill = PatternFill(fill_type=None)
        self.red = PatternFill(fgColor='FFCCCB',
                    fill_type='solid')
        self.red2 = PatternFill(fgColor='F5C2C1',
                    fill_type='solid')
        self.green = PatternFill(fgColor='90EE90',
                    fill_type='solid')
        self.green2 = PatternFill(fgColor='86E486',
                    fill_type='solid')
        self.yellow = PatternFill(fgColor='FFFFE0',
                    fill_type='solid')
        self.yellow2 = PatternFill(fgColor='F5F5D6',
                    fill_type='solid')
        self.gray = PatternFill(fgColor='D3D3D3',
                    fill_type='solid')
        self.gray2 = PatternFill(fgColor='C9C9C9',
                    fill_type='solid')

        try:
            self.wb = load_workbook(fileName)
        except:
            self.wb = Workbook()
            logging.info("Creating new work book")


        try:
            self.capital = self.wb['Raids']
        except:
            self.wb.active.title = 'Raids'
            self.capital = self.wb['Raids']

        try:
            self.war = self.wb["War"]
        except:
            self.wb.create_sheet('War')
            self.war = self.wb["War"]
        
        try:
            self.league = self.wb["League"]
        except:
            self.wb.create_sheet('League')
            self.league = self.wb["League"]
            
    def setUpWarColumnHeaders(self, tag, mapPosition, name, attacks, stars, repeat, date):
        print("Settign up war sheet...")
        self.tagPosW = tag
        self.mapPositionW = mapPosition
        self.namePosW = name
        self.attacksPosW = attacks
        self.starsPosW = stars
        self.repeatPosW = repeat
        self.datePosW = date
        self.warSheetSetUp()

    def setUpLeagueColumnHeaders(self, tag, name, attacks, stars, repeat, datePos):
        print("Setting up league war sheet...")
        self.tagPosL = tag
        self.namePosL = name
        self.attacksPosL = attacks
        self.starsPosL = stars
        self.repeatPosL = repeat

        self.datePosL = datePos

        self.leagueSetUp()


    def setUpRaidColumnHeaders(self, tag, trophies, position, name, attacks, stars, dono, donoR, repeat,
     date, totalGold, totalDono):
        print("Setting up raid sheet...")
        self.tagPosR = tag
        self.trophiesPosR = trophies
        self.positionPosR = position
        self.namePosR = name
        self.attacksPosR = attacks
        self.goldPosR = stars
        self.donationPosR = dono
        self.donationRecievedR = donoR
        self.repeatPosR = repeat
        self.datePosR = date
        self.totalGoldR = totalGold
        self.totalDonoR = totalDono
        self.raidSetUp()

    def leagueSetUp(self):
        self.league.cell(3, self.tagPosL).value = 'Tag'
        self.league.cell(3, self.namePosL).value = 'Name'
        self.league.cell(3, self.attacksPosL).value = 'Attacks'
        self.league.cell(3, self.starsPosL).value = 'Stars'

    def raidSetUp(self):
        self.capital.cell(3, self.tagPosR).value = 'Tag'
        self.capital.cell(3, self.trophiesPosR).value = 'Trophies'
        self.capital.cell(3, self.namePosR).value = 'Name'
        self.capital.cell(3, self.attacksPosR).value = 'Attacks'
        self.capital.cell(3, self.goldPosR).value = 'Gold'
        self.capital.cell(3, self.donationPosR).value = 'D'
        self.capital.cell(3, self.donationRecievedR).value = 'DR'

    def warSheetSetUp(self):
        self.war.cell(3, self.tagPosW).value = 'Tag'
        self.war.cell(3, self.mapPositionW).value = 'Position'
        self.war.cell(3, self.namePosW).value = 'Name'
        self.war.cell(3, self.attacksPosW).value = 'Attacks'
        self.war.cell(3, self.starsPosW).value = 'Stars'

    def setUpMembers(self, clanData):
        totalDonations = 0
        if self.clanMembers == {}:
            tempMembers = clanData['memberList']
            for i, p in enumerate(tempMembers):
                self.clanMembers[p['tag']] = tempMembers[i]
                self.clanMembers[p['tag']]['attacks'] = 0
                self.clanMembers[p['tag']]['capitalResourcesLooted'] = 0
                self.clanMembers[p['tag']]['status'] = 'In'
                self.clanMembers[p['tag']]['attackNumber'] = None
                self.clanMembers[p['tag']]['stars'] = None
                self.clanMembers[p['tag']]['mapPosition'] = None
                totalDonations += tempMembers[i]['donations']

        self.totalDonations = totalDonations
    
    def setUpLeague(self, leagueWarsData, season):
        print("Setting up clan league data...")
        rowMax = self.league.max_row
        colMax = self.league.max_column

        newInfo = False
        self.season = season

        if not season == str(self.league.cell(2, self.datePosL).value)[:7] and not self.league.cell(2, self.datePosL).value == None:
            self.seasons.append(str(self.league.cell(2, self.datePosL).value))
            newInfo = True

        
        for c in range(self.repeatPosL, colMax+1, 2):
            self.seasons.append(str(self.league.cell(2, c).value))

        for r in range(4, rowMax+1):
            tag = self.league.cell(r, self.tagPosL).value
            
            if not tag in self.clanMembers:
                self.clanMembers[tag] = {}
                if newInfo:
                    attackCell = self.league.cell(r, self.attacksPosL).value
                    attackCell = attackCell.split("/")
                    attacks = attackCell[0]
                    maxAttacks = attackCell[1]
                    self.clanMembers[tag][self.seasons[0]] = [
                        attacks,
                        maxAttacks,
                        self.league.cell(r, self.starsPosL).value.split("/")[0]
                    ]
                self.clanMembers[tag]['attacks'] = self.league.cell(r, self.namePosL).value
                self.clanMembers[tag]['name'] = self.league.cell(r, self.namePosL).value
                self.clanMembers[tag]['tag'] = self.league.cell(r, self.tagPosL).value
                self.clanMembers[tag]['status'] = 'Out'
                self.clanMembers[tag]['attackNumber'] = None
                self.clanMembers[tag]['stars'] = None
                self.clanMembers[tag]['mapPosition'] = None
            else:
                if newInfo:
                    attackCell = self.league.cell(r, self.attacksPosL).value
                    if attackCell != None:
                        attackCell = attackCell.split("/")
                        attacks = attackCell[0]
                        maxAttacks = attackCell[1]
                    else:
                        attacks = None
                        maxAttacks = None

                    if self.league.cell(r, self.starsPosL).value != None: 
                        stars = self.league.cell(r, self.starsPosL).value.split("/")[0]
                    else:
                        stars = None

                    self.clanMembers[tag][self.seasons[0]] = [
                        attacks,
                        maxAttacks,
                        stars
                    ]
                
            for c in range(self.repeatPosL, colMax+1, 2):
                date = self.league.cell(2, c).value

                attackCell = self.league.cell(r, c).value
                if attackCell != None:

                    attackCell = attackCell.split("/")
                    attacks = attackCell[0]
                    maxAttacks = attackCell[1]
                else:
                    attacks = None
                    maxAttacks = None
                
                if self.league.cell(r, c+1).value != None: 
                    stars = self.league.cell(r, c+1).value.split("/")[0]
                else:
                    stars = None

                self.clanMembers[tag][date] = [
                    attacks,
                    maxAttacks,
                    stars
                ]
    
        for m in self.clanMembers:
                for d in self.seasons:
                    if not d in self.clanMembers[m]:
                        self.clanMembers[m][d] = [None, None]
                        self.clanMembers[m]['attacks'] = 0
                        self.clanMembers[m]['capitalResourcesLooted'] = None
                        self.clanMembers[m]['status'] = 'In'
                        self.clanMembers[m]['attackNumber'] = None
                        self.clanMembers[m]['stars'] = None
                        self.clanMembers[m]['mapPosition'] = None
        
        
        for warData in leagueWarsData:
            if warData['clan']['tag'] == '#2Y80PGVLJ':
                war = warData['clan']
            else:
                war = warData['opponent']
            for m in war['members']:
                if not m['tag'] in self.clanMembers:
                    self.clanMembers[m['tag']] = {}
                    self.clanMembers[m['tag']]['tag'] = m['tag']
                    
                    self.clanMembers[m['tag']]['attacks'] = 0
                    self.clanMembers[m['tag']]['capitalResourcesLooted'] = None
                    self.clanMembers[m['tag']]['status'] = 'In'
                    self.clanMembers[m['tag']]['attackNumber'] = None
                    self.clanMembers[m['tag']]['stars'] = None
                    self.clanMembers[m['tag']]['mapPosition'] = None

                if not 'maxAttacks' in self.clanMembers[m['tag']]:
                    if not warData['state'] == 'preperation':
                        self.clanMembers[m['tag']]['maxAttacks'] = 1
                else: 
                    if not warData['state'] == 'preparation':
                        self.clanMembers[m['tag']]['maxAttacks'] = self.clanMembers[m['tag']]['maxAttacks'] + 1 
                if 'attacks' in m:
                    if not 'stars' in self.clanMembers[m['tag']]:
                        stars = 0
                    else:
                        stars = self.clanMembers[m['tag']]['stars']
                        if stars == None:
                            stars = 0
                    stars += m['attacks'][0]['stars']
                    if not 'LattackNumber' in self.clanMembers[m['tag']]:
                        self.clanMembers[m['tag']]['LattackNumber'] = len(m['attacks'])
                    else:
                        self.clanMembers[m['tag']]['LattackNumber'] = self.clanMembers[m['tag']]['LattackNumber'] + len(m['attacks'])
                    self.clanMembers[m['tag']]['stars'] = stars
                else: 
                    if not 'LattackNumber' in self.clanMembers[m['tag']]:
                        self.clanMembers[m['tag']]['LattackNumber'] = 0
                        self.clanMembers[m['tag']]['stars'] = 0
                    
    def setUpWar(self, warData):
        print("Settign up war info...")
        rowMax = self.war.max_row
        colMax = self.war.max_column

        
        warD = warData['preparationStartTime'][0:8]
        self.warTime = warD[0:4] + "-" + warD[4:6] + "-" + warD[6:8]
        newInfo = False


        if not self.warTime == str(self.war.cell(2, self.datePosW).value)[:10] and not self.war.cell(2, self.datePosW).value == None:
            self.warDates.append(str(self.war.cell(2, self.datePosW).value)[:10])
            newInfo = True


        for c in range(self.repeatPosW, colMax+1, 2):
            self.warDates.append(str(self.war.cell(2, c).value))

        for r in range(4, rowMax+1):
            tag = self.war.cell(r, self.tagPosW).value
            
            if not tag in self.clanMembers:
                self.clanMembers[tag] = {}
                if newInfo:
                    self.clanMembers[tag][self.warDates[0]] = [
                        self.war.cell(r, self.attacksPosW).value,
                        self.war.cell(r, self.starsPosW).value
                    ]
                self.clanMembers[tag]['name'] = self.war.cell(r, self.namePosW).value
                self.clanMembers[tag]['tag'] = self.war.cell(r, self.tagPosW).value
                # self.clanMembers[tag]['trophies'] = self.war.cell(r, self.mapPositionW).value
                self.clanMembers[tag]['status'] = 'Out'
                self.clanMembers[tag]['attackNumber'] = None
                self.clanMembers[tag]['stars'] = None
                self.clanMembers[tag]['mapPosition'] = None
            else:
                if newInfo:
                    self.clanMembers[tag][self.warDates[0]] = [
                        self.war.cell(r, self.attacksPosW).value,
                        self.war.cell(r, self.starsPosW).value
                    ]
                
            for c in range(self.repeatPosW, colMax+1, 2):
                date = self.war.cell(2, c).value
                self.clanMembers[tag][date] = [
                    self.war.cell(r, c).value,
                    self.war.cell(r, c+1).value
                ]

        for m in self.clanMembers:
            for d in self.warDates:
                if not d in self.clanMembers[m]:
                    self.clanMembers[m][d] = [None, None]
                    self.clanMembers[m]['attacks'] = 0
                    self.clanMembers[m]['capitalResourcesLooted'] = None
                    self.clanMembers[m]['status'] = 'In'
                    self.clanMembers[m]['attackNumber'] = None
                    self.clanMembers[m]['stars'] = None
                    self.clanMembers[m]['mapPosition'] = None

        self.warAmount = 0
        for m in warData['clan']['members']:
            self.warAmount += 1
            if 'attacks' in m:
                stars = 0
                for s in m['attacks']:
                    stars += s['stars']
                self.clanMembers[m['tag']]['attackNumber'] = len(m['attacks'])
                self.clanMembers[m['tag']]['stars'] = stars
                self.clanMembers[m['tag']]['mapPosition'] = m['mapPosition']
            else: 
                self.clanMembers[m['tag']]['attackNumber'] = 0
                self.clanMembers[m['tag']]['stars'] = 0
                self.clanMembers[m['tag']]['mapPosition'] = m['mapPosition']

    def setUpRaids(self, raidData):

        print("Setting up said info...")
        rowMax = self.capital.max_row
        colMax = self.capital.max_column

        state = raidData['items'][0]['state']
        self.totalAttacks = raidData['items'][0]['totalAttacks']
        self.disctrictsDestroyed = raidData['items'][0]['enemyDistrictsDestroyed']
        if self.disctrictsDestroyed != 0:
            self.average = round(self.totalAttacks / self.disctrictsDestroyed, 2)
        else:
            self.average = 0

        self.totalGold = raidData['items'][0]['capitalTotalLoot']

        raidT = raidData['items'][0]['startTime'][0:8]
        self.raidTime = raidT[0:4] + "-" + raidT[4:6] + "-" + raidT[6:8]
        self.raidGolds = []
        self.averages = []
        self.medals = []
        newInfo = False

        if not self.raidTime == str(self.capital.cell(1, self.datePosR).value)[:10] and not self.capital.cell(1, self.datePosR).value == None:
            self.raidDates.append(str(self.capital.cell(1, self.datePosR).value)[:10])
            self.averages.append(self.capital.cell(2, self.datePosR).value)
            self.raidGolds.append(self.capital.cell(2, self.totalGoldR).value)
            self.medals.append(self.capital.cell(1, self.totalGoldR).value)
            self.capital.cell(1, self.totalGoldR).value = "(Medals)"
            newInfo = True

        for c in range(self.repeatPosR, colMax+1, 2):
            self.raidDates.append(str(self.capital.cell(1, c).value))
            self.averages.append(self.capital.cell(2, c).value)
            self.raidGolds.append(self.capital.cell(2, c+1).value)
            self.medals.append(self.capital.cell(1, c+1).value)

        for r in range(4, rowMax+1):
            tag = self.capital.cell(r, self.tagPosR).value

            if not tag in self.clanMembers:
                self.clanMembers[tag] = {}
                self.clanMembers[tag]['tag'] = self.capital.cell(r, self.tagPosR).value
                self.clanMembers[tag]['name'] = self.capital.cell(r, self.namePosR).value
            # self.clanMembers[tag]['trophies'] = self.capital.cell(r, self.mapPositionW).value
                self.clanMembers[tag]['status'] = 'Out'
                self.clanMembers[tag]['donations'] = None
                self.clanMembers[tag]['donationsReceived'] = None
                self.clanMembers[tag]['trophies'] = self.capital.cell(r, self.trophiesPosR).value
                self.clanMembers[tag]['attacks'] = None
                self.clanMembers[tag]['capitalResourcesLooted'] = None
                if newInfo:
                    self.clanMembers[tag][self.raidDates[0]] = [
                        self.capital.cell(r, self.attacksPosR).value,
                        self.capital.cell(r, self.goldPosR).value
                    ]
            else:
                if state == 'ended':
                    flag = True
                    for m in raidData['items'][0]['members']:
                        if m['tag'] == tag:
                            flag = False
                    if flag:
                        self.clanMembers[tag]['capitalResourcesLooted'] = self.capital.cell(r, self.goldPosR).value
                        self.clanMembers[tag]['attacks'] = self.capital.cell(r, self.attacksPosR).value


                if newInfo:
                    self.clanMembers[tag][self.raidDates[0]] = [
                        self.capital.cell(r, self.attacksPosR).value,
                        self.capital.cell(r, self.goldPosR).value
                    ]

            for c in range(self.repeatPosR, colMax+1, 2):
                date = self.capital.cell(1, c).value
                self.clanMembers[tag][date] = [
                    self.capital.cell(r, c).value,
                    self.capital.cell(r, c+1).value
                ]

        for m in self.clanMembers:
            for d in self.raidDates:
                if not d in self.clanMembers[m]:
                    self.clanMembers[m][d] = [None, None]
                    self.clanMembers[m]['attacks'] = None
                    self.clanMembers[m]['capitalResourcesLooted'] = None
                    self.clanMembers[m]['donations'] = None
                    self.clanMembers[m]['donationsReceived'] = None
                    self.clanMembers[m]['status'] = 'In'
                    self.clanMembers[m]['attackNumber'] = None
                    self.clanMembers[m]['stars'] = None
                    self.clanMembers[m]['mapPosition'] = None

        for m in raidData['items'][0]['members']:
            self.clanMembers[m['tag']]['attacks'] = m['attacks']
            self.clanMembers[m['tag']]['capitalResourcesLooted'] = m['capitalResourcesLooted']

    def updateLeagueSheet(self, totalThrshold, attackThreshold, starsThreshold):
        print("Updating war sheet")
        self.updateLeagueSeasons()

        tags = []
        for m in self.clanMembers:
            tags.append(m)

        for r,m in enumerate(tags,4):
            points = -1
            self.writeCell(self.clanMembers[m], r, self.tagPosL, 'tag', self.league, tag=True)
            self.writeCell(self.clanMembers[m], r, self.namePosL, 'name', self.league)

            points += self.writeLeagueAttackCell(self.clanMembers[m], r,self.attacksPosL,0, self.league, params=attackThreshold)
            points += self.writeLeagueStarCell(self.clanMembers[m], r,self.starsPosL,0, self.league, params=starsThreshold)
            
            self.colorName(r, self.namePosL, points, totalThrshold, self.league)
            self.league.cell(r, self.starsPosL).border = self.sideBorder

            for i, d in enumerate(self.seasons):
                c = (((i+1)*2)+(self.repeatPosL-2))
                self.writeLeagueAttackCell(self.clanMembers[m], r,c, d, self.league, params=attackThreshold, dated = True)
                self.writeLeagueStarCell(self.clanMembers[m], r,c+1, d, self.league, params=starsThreshold, dated = True)
                self.league.cell(r, c+1).border = self.sideBorder

    def writeLeagueAttackCell(self, member, r, c, val, sheet, params, dated = False):
        if 'maxAttacks' in member:
            if dated:
                
                a = int(member[val][0])
                m = int(member[val][1])
                vall = str(a) + "/" + str(m)
                p = a / m
            else:
                a = member['LattackNumber']
                m = member['maxAttacks']
                vall = str(a) + "/" + str(m)
                p = a / m
            self.league.cell(r, c).alignment = Alignment(horizontal='center')
            self.league.cell(r, c).value = vall
            if p >= params[0]:
                self.colorSet(self.green, self.green2, r, c, sheet)
                return 3
            elif p >= params[1]:
                self.colorSet(self.yellow, self.yellow2, r, c, sheet)
                return 2
            else :
                self.colorSet(self.red, self.red2, r, c, sheet)
                return 1
        else: 
            self.colorSet(self.gray, self.gray2, r, c, sheet)
            sheet.cell(r, c).value = None
            return 0

    def writeLeagueStarCell(self, member, r, c, val, sheet, params, dated=False):
        if 'maxAttacks' in member:
            something  = True
            if dated:
                vall = str(member[val][2]) + "/" + str(int(member[val][1])*3)
                if member[val][0] == None:
                    something =  True
                else:
                    a = int(member[val][2])
                    m = int(member[val][1])*3
                    p = a / m
            else:
                vall = str(member['stars']) + "/" + str(member['maxAttacks']*3)
                # if member[val][0] == None:
                #     something =  True
                # else:
                a = member['stars']
                m = member['maxAttacks']*3
                p = a / m
            if something:
                self.league.cell(r, c).value = vall
                self.league.cell(r, c).alignment = Alignment(horizontal='center')
                if p >= params[0]:
                    self.colorSet(self.green, self.green2, r, c, sheet)
                    return 3
                elif p >= params[1]:
                    self.colorSet(self.yellow, self.yellow2, r, c, sheet)
                    return 2
                else :
                    self.colorSet(self.red, self.red2, r, c, sheet)
                    return 1
        
        self.colorSet(self.gray, self.gray2, r, c, sheet)
        sheet.cell(r, c).value = None
        return 0

    def updateLeagueSeasons(self):
        self.league.cell(2, self.datePosL).value = self.season
        for i,d in enumerate(self.seasons):
            c = (((i+1)*2)+(self.repeatPosL-2))
            self.league.cell(2, c).value = self.seasons[i]
            self.league.cell(3, c).value = "Attacks"
            self.league.cell(3, c+1).value = "Stars"

    def updateRiadsSheet(self, goldThreshold, attackThreshold, donationsThreshold, totalThreshold):
        print("Updating raid sheet...")
        self.updateRaidVals()

        tags = self.sortGold(self.clanMembers)

        self.capital.column_dimensions['C'].width = 5
        self.capital.column_dimensions['G'].width = 7
        self.capital.column_dimensions['H'].width = 7
        for i in range(1,4):
            self.capital.cell(i, self.donationRecievedR).border = self.sideBorder
            self.capital.cell(i, self.goldPosR).border = self.sideBorder
            self.capital.cell(i, self.namePosR).border = self.sideBorder

        for r,m in enumerate(tags,4):
            colMax = self.capital.max_column
            for c in range(1, colMax+1):
                if (self.clanMembers[m]['attacks'] == None or self.clanMembers[m]['attacks'] == 0) and self.notAttackedFlag == 0:
                    self.capital.cell(r, c).border = self.topBorder

                elif self.clanMembers[m]['status'] == 'Out' and self.clanMembers[m]['capitalResourcesLooted'] == None and self.outFlag == 0:
                    self.capital.cell(r, c).border = self.topBorder

                else:
                    self.capital.cell(r, c).border = None


            if self.clanMembers[m]['status'] == 'Out' and self.clanMembers[m]['capitalResourcesLooted'] == None and self.outFlag == 0:
                self.outFlag = 1

            if (self.clanMembers[m]['attacks'] == None or self.clanMembers[m]['attacks'] == 0) and self.notAttackedFlag == 0:
                self.notAttackedFlag = 1


            points = -1
            self.writeRank(r, self.capital)
            self.writeCell(self.clanMembers[m], r, self.tagPosR, 'tag', self.capital, tag=True)
            self.writeCell(self.clanMembers[m], r, self.namePosR, 'name', self.capital)
            self.writeCell(self.clanMembers[m], r, self.trophiesPosR, 'trophies', self.capital)
            points += self.writeCell(self.clanMembers[m], r, self.attacksPosR, 'attacks', self.capital, params=attackThreshold)
            points += self.writeCell(self.clanMembers[m], r, self.goldPosR, 'capitalResourcesLooted', self.capital, params=goldThreshold)
            self.writeCell(self.clanMembers[m], r, self.donationPosR, 'donations', self.capital, params=donationsThreshold)
            self.writeCell(self.clanMembers[m], r, self.donationRecievedR, 'donationsReceived', self.capital)
            self.colorName(r, self.namePosR, points, totalThreshold, self.capital)

            if  self.outFlag == 1 or self.notAttackedFlag == 1:
                self.capital.cell(r, self.donationRecievedR).border = self.topNSideBorder
                self.capital.cell(r, self.goldPosR).border = self.topNSideBorder
                self.capital.cell(r, self.namePosR).border = self.topNSideBorder
            else:
                self.capital.cell(r, self.donationRecievedR).border = self.sideBorder
                self.capital.cell(r, self.goldPosR).border = self.sideBorder
                self.capital.cell(r, self.namePosR).border = self.sideBorder

            self.underLineName(self.clanMembers[m], r, self.namePosR, self.capital)
            
            
            for i, d in enumerate(self.raidDates):
                c = (((i+1)*2)+(self.repeatPosR-2))

                for g in range(1,4):
                    self.capital.cell(g, c+1).border = self.sideBorder

                self.writeCell(self.clanMembers[m], r,c, d, self.capital, params=attackThreshold, dated = 0)
                self.writeCell(self.clanMembers[m], r,c+1, d, self.capital, params=goldThreshold, dated = 1)
                if  self.outFlag == 1 or self.notAttackedFlag == 1:
                    self.capital.cell(r, c+1).border = self.topNSideBorder
                else:
                    self.capital.cell(r, c+1).border = self.sideBorder


            if self.outFlag == 1:
                self.outFlag = 2

            if self.notAttackedFlag == 1:
                self.notAttackedFlag = 2

    def writeRank(self, r, sheet):
        sheet.cell(r, self.positionPosR).value = r - 3
        sheet.cell(r, self.positionPosR).alignment = Alignment(horizontal='center')
        self.colorSet(self.gray, self.gray2, r, self.positionPosR, self.capital)

    def updateRaidVals(self):
        tick = 0
        try:
            t = (int)(self.capital.cell(1, self.totalGoldR).value)
            if t > self.medals[0]:
                self.colorSet(self.green, self.green, 1, self.totalGoldR, self.capital)
            else:
                self.colorSet(self.red, self.red, 1, self.totalGoldR, self.capital)
                tick += 1
        except:
            pass

        self.capital.cell(2, self.totalGoldR).value = self.totalGold
        if self.totalGold > self.raidGolds[0]:
            self.colorSet(self.green, self.green, 2, self.totalGoldR, self.capital)
        else:
            self.colorSet(self.red, self.red, 2, self.totalGoldR, self.capital)
            tick += 1


        self.capital.cell(2, self.totalGoldR).number_format  = self.numFormat

        self.capital.cell(2, self.totalDonoR).value = self.totalDonations
        self.capital.cell(2, self.totalDonoR).number_format  = self.numFormat


        self.capital.cell(2, self.datePosR).value = self.average
        if self.average < self.averages[0]:
            self.colorSet(self.green, self.green, 2, self.datePosR, self.capital, False)
        else:
            self.colorSet(self.red, self.red, 2, self.datePosR, self.capital, False)
            tick += 1


        self.capital.cell(1, self.datePosR).value = self.raidTime
        if tick < 2:
            self.colorSet(self.green, self.green, 1, self.datePosR, self.capital)
        else:
            self.colorSet(self.red, self.red, 1, self.datePosR, self.capital)

        for i,d in enumerate(self.raidDates):
            tick = 0
            c = (((i+1)*2)+(self.repeatPosR-2))
            self.capital.cell(2, c).value = self.averages[i]
            self.capital.cell(2, c+1).value = self.raidGolds[i]
            self.capital.cell(1, c+1).value = self.medals[i]
            self.capital.cell(1, c).value = self.raidDates[i]
            self.capital.cell(3, c).value = "Attacks"
            self.capital.cell(3, c+1).value = "Gold"

            if i == len(self.raidDates) -1:
                continue

            if self.averages[i] < self.averages[i+1]:
                self.colorSet(self.green, self.green, 2, c, self.capital, False)
            else:
                self.colorSet(self.red, self.red, 2, c, self.capital, False)
                tick += 1

            if self.raidGolds[i] > self.raidGolds[i+1]:
                self.colorSet(self.green, self.green, 2, c+1, self.capital)
            else:
                self.colorSet(self.red, self.red, 2, c+1, self.capital)
                tick += 1

            if self.medals[i] > self.medals[i+1]:
                self.colorSet(self.green, self.green, 1, c+1, self.capital)
            else:
                self.colorSet(self.red, self.red, 1, c+1, self.capital)
                tick += 1

            if tick < 2:
                self.colorSet(self.green, self.green, 1, c, self.capital)
            else:
                self.colorSet(self.red, self.red, 1, c, self.capital)


    def sortGold(self, members):
        temp = []

        self.raidDates

        for i, p in enumerate(members):
            temp.append(p)

        for i, p in enumerate(temp):
            best = i
            for j in range(i+1,len(temp)):
                if i == j:
                    continue
                if members[temp[j]]['status'] == 'Out' and members[temp[j]]['capitalResourcesLooted'] == None:
                    continue
                elif members[temp[j]]['capitalResourcesLooted'] == None:
                    if members[temp[best]]['capitalResourcesLooted'] == 0:
                        best = j
                        pass
                    continue
                elif members[temp[best]]['capitalResourcesLooted'] == None:
                    if members[temp[j]]['capitalResourcesLooted'] != 0:
                        best = j

                elif members[temp[best]]['capitalResourcesLooted'] == 0 and members[temp[j]]['capitalResourcesLooted'] == 0:
                    for d in self.raidDates: 
                        if members[temp[j]][d][1] == None and members[temp[best]][d][1] != None:
                            best = j
                            break
                        elif members[temp[j]][d][1] == None:
                            continue
                        elif members[temp[best]][d][1] == None:
                            # best = j
                            continue
                        elif members[temp[best]][d][1] == 0 and members[temp[j]][d][1] == 0:
                            continue
                        elif members[temp[best]][d][1] < members[temp[j]][d][1]:
                            best = j
                            break
                        elif members[temp[best]][d][1] == members[temp[j]][d][1]:
                            if members[temp[best]]['donations'] < members[temp[j]]['donations']:
                                best = j
                            break
                        elif members[temp[best]][d][1] > members[temp[j]][d][1]:
                            break

                elif members[temp[best]]['capitalResourcesLooted'] < members[temp[j]]['capitalResourcesLooted']:
                    best = j
                elif members[temp[best]]['capitalResourcesLooted'] == members[temp[j]]['capitalResourcesLooted']:
                    if members[temp[best]]['donations'] < members[temp[j]]['donations']:
                        best = j
                    continue

            tt = temp[i]
            temp[i] = temp[best]
            temp[best] = tt

        return temp

    def sortPositino(self, members):
        temp = [None] * self.warAmount
        trash = []

        for m in members:
            if not members[m]['mapPosition'] == None:
                temp[members[m]['mapPosition']-1] = m
            else:
                trash.append(m)

        return temp + trash

    def updateWarSheet(self, attackThreshold, starsThreshold, totalThrshold):
        print("Updating war sheet")
        self.updateWarTimes()

        tags = self.sortPositino(self.clanMembers)

        for r,m in enumerate(tags,4):
            points = -1
            self.writeCell(self.clanMembers[m], r, self.tagPosW, 'tag', self.war, tag=True)
            self.writeCell(self.clanMembers[m], r, self.mapPositionW, 'mapPosition', self.war)
            self.writeCell(self.clanMembers[m], r, self.namePosW, 'name', self.war)
            points += self.writeCell(self.clanMembers[m], r, self.attacksPosW, 'attackNumber', self.war, params=attackThreshold)
            points += self.writeCell(self.clanMembers[m], r, self.starsPosW, 'stars', self.war, params=starsThreshold)
            self.colorName(r, self.namePosW, points, totalThrshold, self.war)
            self.war.cell(r, self.starsPosW).border = self.sideBorder

            for i, d in enumerate(self.warDates):
                c = (((i+1)*2)+(self.repeatPosW-2))
                self.writeCell(self.clanMembers[m], r,c, d, self.war, params=attackThreshold, dated = 0)
                self.writeCell(self.clanMembers[m], r,c+1, d, self.war, params=starsThreshold, dated = 1)
                self.war.cell(r, c+1).border = self.sideBorder

    def updateWarTimes(self):
        self.war.cell(2, self.datePosW).value = self.warTime
        for i,d in enumerate(self.warDates):
            c = (((i+1)*2)+(self.repeatPosW-2))
            self.war.cell(2, c).value = self.warDates[i]
            self.war.cell(3, c).value = "Attacks"
            self.war.cell(3, c+1).value = "Stars"

    def colorName(self, r, c, points, params, sheet):
        if points >= params[0]:
            self.colorSet(self.green, self.green2, r, c, sheet)
        elif points >= params[1]:
            self.colorSet(self.yellow, self.yellow2, r, c, sheet)
        elif not points == -1:
            self.colorSet(self.red, self.red2, r, c, sheet)

    def writeCell(self, member, r, c, val, sheet, params=None, tag=False, dated=-1):
        if member['tag'] in self.tags:
            # sheet.cell(r, c).border = thin_border
            sheet.cell(r, c).font = Font(bold=True)

        else:
            sheet.cell(r, c).font = Font(bold=False)

        if self.outFlag == 1:
            sheet.cell(r, c).border = self.topBorder

        elif self.notAttackedFlag == 1:
            sheet.cell(r, c).border = self.topBorder
        else:
            sheet.cell(r, c).border = None

        # if (r-3) % 11 == 0:
        #     sheet.cell(r, c).border = self.thin_border

        if tag == True:
            sheet.cell(r, c).value = member[val]
            if member[val] in self.tags:
                sheet.cell(r, c).font = Font(bold=True)
            else:
                sheet.cell(r, c).font = Font(bold=False)
            if member['status'] == 'In':
                self.colorSet(self.gray, self.gray2, r, c, sheet)
            else:
                self.colorSet(self.red, self.red2, r, c, sheet)
        elif params == None:
            if val in member:
                sheet.cell(r, c).value = member[val]
                self.colorSet(self.gray, self.gray2, r, c, sheet)
        else:
            if not dated == -1:
                sheet.cell(r, c).value = member[val][dated]
                if member[val][dated] == None:
                    self.colorSet(self.gray, self.gray2, r, c, sheet)
                    return 
                if member[val][dated] >= params[0]:
                    self.colorSet(self.green, self.green2, r, c, sheet)
                    return 3
                elif member[val][dated] >= params[1]:
                    self.colorSet(self.yellow, self.yellow2, r, c, sheet)
                    return 2
                else :
                    self.colorSet(self.red, self.red2, r, c, sheet)
                    return 1
            elif not member[val] == None and val in member: 
                sheet.cell(r, c).value = member[val]
                if member[val] >= params[0]:
                    self.colorSet(self.green, self.green2, r, c, sheet)
                    return 3
                elif member[val] >= params[1]:
                    self.colorSet(self.yellow, self.yellow2, r, c, sheet)
                    return 2
                else :
                    self.colorSet(self.red, self.red2, r, c, sheet)
                    return 1
            else: 
                self.colorSet(self.gray, self.gray2, r, c, sheet)
                sheet.cell(r, c).value = None
                return 0

    def underLineName(self, member, r, c, sheet):
        if(self.notAttackedFlag == 0) :
            if(len(self.raidDates) > 0):
                if(not self.raidDates[0] in member):
                    sheet.cell(r,c).border = self.bottomNSideBorder
                elif(member[self.raidDates[0]][0] == None or member[self.raidDates[0]][0] == 0):
                    sheet.cell(r,c).border = self.bottomNSideBorder

    def colorSet(self, color, color2, r, c, sheet, sameFormat = True):
        if r % 2 == 0:
            sheet.cell(r, c).fill = color
        else:
            sheet.cell(r, c).fill = color2
        if sameFormat:
            sheet.cell(r, c).number_format  = self.numFormat
